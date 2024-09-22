# 우선 이미지고르면 이미 폴더가 보이겠끔 됬어
import os
import sys
import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit, QFileDialog, QScrollArea, QMessageBox, QListWidget
from PyQt5.QtWidgets import QProgressBar
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtCore import Qt

class ImageEvaluatorGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        # 입력 경로
        input_layout = QHBoxLayout()
        self.input_path = QLineEdit()
        input_button = QPushButton('입력 폴더 선택')
        input_button.clicked.connect(self.select_input_folder)
        input_layout.addWidget(QLabel('입력 경로:'))
        input_layout.addWidget(self.input_path)
        input_layout.addWidget(input_button)
        layout.addLayout(input_layout)

        # xlsx 파일 선택
        xlsx_layout = QHBoxLayout()
        self.xlsx_path = QLineEdit()
        xlsx_button = QPushButton('xlsx 파일 선택')
        xlsx_button.clicked.connect(self.select_xlsx_file)
        xlsx_layout.addWidget(QLabel('xlsx 파일:'))
        xlsx_layout.addWidget(self.xlsx_path)
        xlsx_layout.addWidget(xlsx_button)
        layout.addLayout(xlsx_layout)

        # 폴더 목록
        self.folder_list = QListWidget()
        layout.addWidget(QLabel('처리할 폴더:'))
        layout.addWidget(self.folder_list)

        # 출력 경로
        output_layout = QHBoxLayout()
        self.output_path = QLineEdit()
        output_button = QPushButton('이미지 출력 폴더 선택')
        output_button.clicked.connect(self.select_output_folder)
        output_layout.addWidget(QLabel('출력 경로:'))
        output_layout.addWidget(self.output_path)
        output_layout.addWidget(output_button)
        layout.addLayout(output_layout)

        # 밝기 설정
        brightness_layout = QHBoxLayout()
        self.brightness_min = QLineEdit('58')
        self.brightness_max = QLineEdit('200')
        brightness_layout.addWidget(QLabel('밝기 범위:'))
        brightness_layout.addWidget(self.brightness_min)
        brightness_layout.addWidget(QLabel('-'))
        brightness_layout.addWidget(self.brightness_max)
        layout.addLayout(brightness_layout)

        # 선명도 설정
        sharpness_layout = QHBoxLayout()
        self.sharpness_threshold = QLineEdit('50')
        sharpness_layout.addWidget(QLabel('선명도 임계값:'))
        sharpness_layout.addWidget(self.sharpness_threshold)
        layout.addLayout(sharpness_layout)

        # 대비 설정
        contrast_layout = QHBoxLayout()
        self.contrast_min = QLineEdit('40')
        self.contrast_max = QLineEdit('80')
        contrast_layout.addWidget(QLabel('대비 범위:'))
        contrast_layout.addWidget(self.contrast_min)
        contrast_layout.addWidget(QLabel('-'))
        contrast_layout.addWidget(self.contrast_max)
        layout.addLayout(contrast_layout)

        # 어두움 설정
        darkness_layout = QHBoxLayout()
        self.dark_threshold = QLineEdit('40')
        self.dark_ratio = QLineEdit('0.9')
        darkness_layout.addWidget(QLabel('어두움 임계값:'))
        darkness_layout.addWidget(self.dark_threshold)
        darkness_layout.addWidget(QLabel('어두움 비율:'))
        darkness_layout.addWidget(self.dark_ratio)
        layout.addLayout(darkness_layout)

        # 이미지 표시 영역을 프로그레스 바로 대체
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.progress_bar)

        # 처리 버튼
        process_button = QPushButton('이미지 처리')
        process_button.clicked.connect(self.process_folders)
        layout.addWidget(process_button)

        self.setLayout(layout)
        self.setWindowTitle('이미지 평가기')
        self.setGeometry(300, 300, 600, 400)

    def select_input_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "입력 폴더 선택")
        if folder:
            self.input_path.setText(folder)
            self.update_folder_list()

    def select_xlsx_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "xlsx 파일 선택", "", "Excel Files (*.xlsx)")
        if file:
            self.xlsx_path.setText(file)
            self.load_excel_data()
    
    def load_excel_data(self):
        try:
            self.df = pd.read_excel(self.xlsx_path.text(), sheet_name='raw')  # 'saheet_name'을 'sheet_name'으로 수정
            print("Excel 파일 로드 완료")
        except Exception as e:
            print(f"Excel 파일 로드 중 오류 발생: {str(e)}")
        

    def update_folder_list(self):
        self.folder_list.clear()
        base_path = Path(self.input_path.text())
        if base_path.exists() and base_path.is_dir():
            for folder in base_path.iterdir():
                if folder.is_dir():
                    self.folder_list.addItem(folder.name)

    def select_output_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "출력 폴더 선택")
        if folder:
            self.output_path.setText(folder)

    def show_warning(self, message):
        QMessageBox.warning(self, "경고", message, QMessageBox.Ok)

    @staticmethod
    def normalize_filename(filename):
        return filename.replace(':', '_')

    def process_folders(self):
        base_path = Path(self.input_path.text())
        output_base = Path(self.output_path.text())
        if not base_path.exists() or not base_path.is_dir():
            self.show_warning("입력 경로가 존재하지 않거나 디렉토리가 아닙니다.")
            return
        if not output_base.exists() or not output_base.is_dir():
            self.show_warning("출력 경로가 존재하지 않거나 디렉토리가 아닙니다.")
            return

        # 모든 폴더를 처리합니다
        folders = [item.text() for item in self.folder_list.findItems('*', Qt.MatchWildcard)]
        
        if not folders:
            self.show_warning("처리할 폴더가 없습니다.")
            return

        self.results = {}
        num_folders = len(folders)
        for i, folder in enumerate(folders, 1):
            folder_path = base_path / folder
            self.process_folder(folder_path, output_base)
            progress = int((i / num_folders) * 100)
            self.progress_bar.setValue(progress)
            QApplication.processEvents()  # UI 업데이트를 위해 이벤트 처리

        self.update_excel_with_results()
        QMessageBox.information(self, "완료", "모든 이미지 처리가 완료되었습니다.", QMessageBox.Ok)
        
    def process_folder(self, folder_path, output_base):
        for image_file in folder_path.glob('**/*.*'):
            if image_file.suffix.lower() in ['.jpg', '.jpeg', '.png']:
                cust_id = image_file.parent.name
                file_name = image_file.name
                img, result = self.evaluate_image(image_file)
                if img is None:
                    print(result)
                    continue
                
                img_with_text = self.add_text_to_image(img, result, cust_id, file_name)
                
                relative_path = image_file.relative_to(folder_path.parent)
                output_path = output_base / relative_path
                output_path.parent.mkdir(parents=True, exist_ok=True)
                print(f'Attempting to save to: {output_path}')
                
                try:
                    pil_img = Image.fromarray(cv2.cvtColor(img_with_text, cv2.COLOR_BGR2RGB))
                    pil_img.save(str(output_path))
                    print(f"평가된 이미지 저장됨: {output_path}")
                    
                    # # 결과 저장 (cust_id와 file_name을 키로 사용)
                    # if "필수검사" in result:
                    #     self.results[(cust_id, file_name)] = "필수검사"
                    # elif "통과" in result:
                    #     self.results[(cust_id, file_name)] = "통과"
                    # else:
                    #     self.results[(cust_id, file_name)] = "확인필요"

                    if output_path.exists():
                        print(f"파일이 성공적으로 생성됨: {output_path}")
                    else:
                        print(f"파일이 생성되지 않음: {output_path}")
                except Exception as e:
                    print(f"이미지 저장 중 오류 발생: {output_path}")
                    print(f"오류 내용: {str(e)}")

    def update_excel_with_results(self):
        base_path = Path(self.input_path.text())
        output_base = Path(self.output_path.text())
        if not base_path.exists() or not base_path.is_dir():
            self.show_warning("입력 경로가 존재하지 않거나 디렉토리가 아닙니다.")
            return
        if not output_base.exists() or not output_base.is_dir():
            self.show_warning("출력 경로가 존재하지 않거나 디렉토리가 아닙니다.")
            return        
        try:
            # 엑셀 파일 열기
            book = load_workbook(self.xlsx_path.text())
            sheet = book['raw']

            # 열 인덱스 찾기
            header_row = sheet[1]
            file_path_col = None
            file_name_col = None
            result_col = None
            image_path_col = None

            for idx, cell in enumerate(header_row, 1):
                if cell.value == 'file_path':
                    file_path_col = idx
                elif cell.value == 'file_name':
                    file_name_col = idx
                elif cell.value == '평가결과':
                    result_col = idx
                elif cell.value == '이미지경로':
                    image_path_col = idx

            if file_path_col is None or file_name_col is None:
                raise ValueError("'file_path' 또는 'file_name' 열을 찾을 수 없습니다.")

            # '평가결과' 열이 없으면 새로 만들기
            if result_col is None:
                result_col = len(header_row) + 1
                sheet.cell(row=1, column=result_col, value='평가결과')

            # '이미지경로' 열이 없으면 새로 만들기
            if image_path_col is None:
                image_path_col = len(header_row) + 2
                sheet.cell(row=1, column=image_path_col, value='이미지경로')

            # 결과 업데이트
            for row in range(2, sheet.max_row + 1):
                file_path = sheet.cell(row=row, column=file_path_col).value
                file_name = sheet.cell(row=row, column=file_name_col).value
                
                if file_path and file_name:
                    file_path = file_path.strip()
                    file_name = file_name.strip().replace(':', '_')
                    cust_id = file_path.split('/')[-2]
                    
                    # 결과 찾기
                    result = "평가되지 않음"
                    image_path = ""
                    for result_path, eval_result in self.results.items():
                        if cust_id in result_path[0] and file_name in result_path[1]:
                            result = eval_result
                            image_path = str(output_base / Path(file_path) / file_name)
                            break

                    sheet.cell(row=row, column=result_col, value=result)
                    sheet.cell(row=row, column=image_path_col, value=image_path)

            # 변경사항 저장
            book.save(self.xlsx_path.text())
            print("Excel 파일 업데이트 완료")
        except Exception as e:
            print(f"Excel 파일 업데이트 중 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()

    def evaluate_image(self, image_path):
        
        # img = cv2.imread(str(image_path))
        img_array = np.fromfile(str(image_path), np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_UNCHANGED)
        
        
        if img is None:
            return None, f"이미지를 읽을 수 없습니다: {image_path}"
        
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        brightness = np.mean(gray)
        laplacian = cv2.Laplacian(gray, cv2.CV_64F)
        sharpness = np.var(laplacian)
        contrast = gray.std()
        
        # GUI에서 설정한 값 사용
        brightness_min = float(self.brightness_min.text())
        brightness_max = float(self.brightness_max.text())
        sharpness_threshold = float(self.sharpness_threshold.text())
        contrast_min = float(self.contrast_min.text())
        contrast_max = float(self.contrast_max.text())
        dark_threshold = float(self.dark_threshold.text())
        dark_ratio_threshold = float(self.dark_ratio.text())
        
        # 평가
        brightness_status = "적정" if brightness_min < brightness < brightness_max else "부적절"
        sharpness_status = "선명" if sharpness > sharpness_threshold else "흐림"
        contrast_status = "적정" if contrast_min < contrast < contrast_max else "부적절"
        
        # 전체적인 어두움 검사
        dark_pixel_ratio = np.sum(gray < dark_threshold) / gray.size
        darkness_status = "정상" if dark_pixel_ratio < dark_ratio_threshold else "너무 어두움"
        
        result = f"밝기: {brightness:.2f} ({brightness_status})\n"
        result += f"선명도: {sharpness:.2f} ({sharpness_status})\n"
        result += f"대비: {contrast:.2f} ({contrast_status})\n"
        result += f"어두움 비율: {dark_pixel_ratio:.2f} ({darkness_status})"
        
        return img, result
    
    def add_text_to_image(self, image, text,  cust_id, file_name):
        image_pil = Image.fromarray(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
        
        image_height = image.shape[0]
        bar_height = max(150, int(image_height * 0.3))
        font_size = max(20, int(image_height / 30))

        # 평가 결과 확인
        is_too_dark = any("너무 어두움" in line for line in text.split('\n'))
        is_pass = not any("부적절" in line or "흐림" in line or "너무 어두움" in line for line in text.split('\n'))
        
        if is_too_dark:
            bar_color = (255, 0, 0)  # 빨간색 
            pass_text = "필수검사"
            self.results[(cust_id, file_name)] = "필수검사"
        elif is_pass:
            bar_color = (0, 0, 255)  # 파란색
            pass_text = "통과"
            self.results[(cust_id, file_name)] = "통과"
        else: 
            bar_color = (255, 165, 0)  # 주황색 (RGB) 
            pass_text = "확인필요"
            self.results[(cust_id, file_name)] = "확인필요"

        bar = Image.new('RGB', (image.shape[1], bar_height), color=bar_color)
        image_with_bar = Image.new('RGB', (image.shape[1], image.shape[0] + bar_height))
        image_with_bar.paste(bar, (0, 0))
        image_with_bar.paste(image_pil, (0, bar_height))

        draw = ImageDraw.Draw(image_with_bar)
        
        font_path = r"C:\Windows\Fonts\malgun.ttf"
        try:
            font = ImageFont.truetype(font_path, font_size)
        except IOError:
            print(f"폰트를 찾을 수 없습니다: {font_path}. 기본 폰트를 사용합니다.")
            font = ImageFont.load_default()

        y0 = int(bar_height * 0.1)
        dy = int(bar_height * 0.18)

        # 통과 여부 텍스트 추가
        draw.text((10, y0), pass_text, font=font, fill=(255, 255, 255))

        # 평가 결과 텍스트 추가
        for i, line in enumerate(text.split('\n'), start=1):
            y = y0 + i*dy
            draw.text((10, y), line, font=font, fill=(255, 255, 255))

        return cv2.cvtColor(np.array(image_with_bar), cv2.COLOR_RGB2BGR)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ImageEvaluatorGUI()
    ex.show()
    sys.exit(app.exec_())